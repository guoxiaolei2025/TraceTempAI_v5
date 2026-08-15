import io
import logging
from typing import Dict, Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.bar_chart import BarChart

logger = logging.getLogger(__name__)


def _apply_sheet_formatting(ws, header_font_bold=True):
    """为工作表应用统一的边框和列宽自适应格式"""
    thin = Side(border_style='thin', color='000000')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    header_font = Font(bold=header_font_bold, size=11)

    for row in ws.iter_rows():
        for cell in row:
            cell.border = border
            cell.alignment = center_align
            if cell.row == 1:
                cell.font = header_font

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def _add_pie_chart(ws, title, category_col, value_col):
    """在报警类型汇总表中添加饼图"""
    chart = PieChart()
    chart.title = title
    labels = Reference(ws, min_col=category_col, min_row=2, max_row=ws.max_row)
    data = Reference(ws, min_col=value_col, min_row=1, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    ws.add_chart(chart, f"E2")


def _add_stacked_bar_chart(ws, title, data_start_col):
    """为分布表添加堆叠柱状图"""
    chart = BarChart()
    chart.type = "col"
    chart.grouping = "stacked"
    chart.title = title
    chart.y_axis.title = '报警次数'

    data = Reference(ws, min_col=data_start_col, min_row=1,
                     max_col=ws.max_column, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, f"E2")


def generate_trend_excel(trend_data: Dict[str, Any]) -> bytes:
    """
    根据趋势分析数据生成 Excel 文件，包含完整边框和图表。

    Args:
        trend_data: 包含以下键的字典
            - dataframe: pandas DataFrame（趋势分析主数据）
            - timestamp: 时间戳字符串
            - department: 学科名称
            - time_range: 查询时间范围

    Returns:
        bytes: Excel 文件的字节内容
    """
    try:
        df = trend_data.get('dataframe')
        if df is None or df.empty:
            logger.info("趋势分析数据为空，返回空Excel")
            buffer = io.BytesIO()
            wb = Workbook()
            ws = wb.active
            ws.title = '无数据'
            ws['A1'] = '暂无数据'
            _apply_sheet_formatting(ws)
            wb.save(buffer)
            return buffer.getvalue()

        logger.info("开始生成趋势分析Excel，数据行数: %d", len(df))
        buffer = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)

        # Sheet 1: 原始报警数据
        raw_df = df.copy()
        display_cols = [
            'device_group', 'device_sn', 'alarmdate', 'message',
            'alarm_type', 'handlestate', 'handleremark', 'remark',
            'alarmtrigger'
        ]
        available_cols = [c for c in display_cols if c in raw_df.columns]
        ws_raw = wb.create_sheet('原始报警数据')
        for row in dataframe_to_rows(raw_df[available_cols], index=False, header=True):
            ws_raw.append(row)
        _apply_sheet_formatting(ws_raw)

        # Sheet 2: 设备报警汇总
        if 'device_sn' in df.columns:
            device_summary = df['device_sn'].value_counts().reset_index()
            device_summary.columns = ['设备编号', '报警次数']
            ws_device = wb.create_sheet('设备报警汇总')
            for row in dataframe_to_rows(device_summary, index=False, header=True):
                ws_device.append(row)
            _apply_sheet_formatting(ws_device)
            _add_stacked_bar_chart(ws_device, '设备报警次数', 2)

        # Sheet 3: 报警类型汇总
        if 'alarm_type' in df.columns:
            type_summary = df['alarm_type'].value_counts().reset_index()
            type_summary.columns = ['报警类型', '报警次数']
            ws_type = wb.create_sheet('报警类型汇总')
            for row in dataframe_to_rows(type_summary, index=False, header=True):
                ws_type.append(row)
            _apply_sheet_formatting(ws_type)
            _add_pie_chart(ws_type, '报警类型占比', 1, 2)

        # Sheet 4: 24小时分布
        if 'hour' in df.columns and 'alarm_type' in df.columns:
            hour_summary = df.groupby(['hour', 'alarm_type']).size().unstack(fill_value=0)
            ws_hour = wb.create_sheet('24小时分布')
            for row in dataframe_to_rows(hour_summary, index=True, header=True):
                ws_hour.append(row)
            _apply_sheet_formatting(ws_hour)
            _add_stacked_bar_chart(ws_hour, '24小时报警分布', 2)

        # Sheet 5: 星期分布
        if 'day_of_week' in df.columns and 'alarm_type' in df.columns:
            day_order = ['Monday', 'Tuesday', 'Wednesday', 'Thursday',
                         'Friday', 'Saturday', 'Sunday']
            day_summary = df.groupby(['day_of_week', 'alarm_type']).size().unstack(fill_value=0)
            day_summary = day_summary.reindex(
                [d for d in day_order if d in day_summary.index], fill_value=0
            )
            ws_day = wb.create_sheet('星期分布')
            for row in dataframe_to_rows(day_summary, index=True, header=True):
                ws_day.append(row)
            _apply_sheet_formatting(ws_day)
            _add_stacked_bar_chart(ws_day, '星期报警分布', 2)

        # Sheet 6: 月度日期分布
        if 'day_of_month' in df.columns and 'alarm_type' in df.columns:
            month_summary = df.groupby(['day_of_month', 'alarm_type']).size().unstack(fill_value=0)
            ws_month = wb.create_sheet('月度日期分布')
            for row in dataframe_to_rows(month_summary, index=True, header=True):
                ws_month.append(row)
            _apply_sheet_formatting(ws_month)
            _add_stacked_bar_chart(ws_month, '月度日期报警分布', 2)

        wb.save(buffer)
        logger.info("趋势分析Excel生成完成")
        return buffer.getvalue()
    except Exception as e:
        logger.exception("生成趋势分析Excel失败: %s", e)
        raise
